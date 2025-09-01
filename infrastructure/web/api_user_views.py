"""
API Views para operaciones de usuario usando Clean Architecture.
"""
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes, authentication_classes
from rest_framework.permissions import IsAuthenticated, IsAdminUser
from rest_framework.response import Response
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from rest_framework.authentication import TokenAuthentication
from rest_framework_simplejwt.authentication import JWTAuthentication
import json

from infrastructure.dependency_setup import get_user_service
from application.dtos import (
    CreateUserRequest,
    UpdateUserRequest,
    LoginRequest,
    ChangePasswordRequest,
)
from .utils import (
    create_api_response,
    handle_api_exception,
    serialize_user_data,
    create_pagination_data,
    validate_pagination_params
)


@api_view(['GET'])
def user_list(request):
    """
    GET /api/v1/users/
    Listar usuarios con paginación y filtros.
    """
    try:
        user_service = get_user_service()
        
        # Validar parámetros de paginación
        pagination_params = validate_pagination_params(request)
        
        # Otros parámetros de filtro
        role_filter = request.GET.get('role')
        active_filter = request.GET.get('active')
        
        # Convertir active_filter a boolean si está presente
        if active_filter is not None:
            active_filter = active_filter.lower() == 'true'
        
        # Llamar al servicio
        result = user_service.list_users(
            page=pagination_params['page'],
            page_size=pagination_params['page_size'],
            role_filter=role_filter,
            active_filter=active_filter
        )
        
        # Serializar datos
        users_data = [serialize_user_data(user) for user in result.users]
        
        return create_api_response(
            data={
                'users': users_data,
                'pagination': create_pagination_data(result)
            }
        )
        
    except Exception as e:
        return handle_api_exception(e, "Error al listar usuarios")


@api_view(['GET'])
def user_detail(request, user_id):
    """
    GET /api/v1/users/{id}/
    Obtener detalles de un usuario específico.
    """
    try:
        user_service = get_user_service()
        
        user = user_service.get_user(user_id)
        
        if not user:
            return create_api_response(
                success=False,
                error='Usuario no encontrado',
                status_code=status.HTTP_404_NOT_FOUND
            )
        
        return create_api_response(
            data={'user': serialize_user_data(user)}
        )
        
    except Exception as e:
        return handle_api_exception(e, "Error al obtener detalles del usuario")


@api_view(['POST'])
@permission_classes([IsAdminUser])
def user_create(request):
    """
    POST /api/v1/users/
    Crear un nuevo usuario.
    """
    try:
        user_service = get_user_service()
        
        # Validar datos requeridos
        required_fields = ['username', 'email', 'full_name', 'password']
        for field in required_fields:
            if field not in request.data:
                return Response({
                    'success': False,
                    'error': f'Campo requerido: {field}'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        # Crear DTO de solicitud
        create_request = CreateUserRequest(
            username=request.data['username'],
            email=request.data['email'],
            full_name=request.data['full_name'],
            password=request.data['password'],
            role=request.data.get('role', 'user'),
            is_active=request.data.get('is_active', True)
        )
        
        # Crear usuario
        created_user = user_service.create_user(create_request)
        
        return Response({
            'success': True,
            'data': {
                'id': created_user.id,
                'username': created_user.username,
                'email': created_user.email,
                'full_name': created_user.full_name,
                'role': created_user.role,
                'is_active': created_user.is_active,
            },
            'message': 'Usuario creado exitosamente'
        }, status=status.HTTP_201_CREATED)
        
    except ValueError as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def user_update(request, user_id):
    """
    PUT /api/v1/users/{id}/
    Actualizar un usuario existente.
    """
    try:
        user_service = get_user_service()
        
        # Verificar que el usuario existe
        existing_user = user_service.get_user(user_id)
        if not existing_user:
            return Response({
                'success': False,
                'error': 'Usuario no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Crear DTO de actualización
        update_request = UpdateUserRequest(
            full_name=request.data.get('full_name'),
            email=request.data.get('email'),
            role=request.data.get('role'),
            is_active=request.data.get('is_active')
        )
        
        # Actualizar usuario
        updated_user = user_service.update_user(user_id, update_request)
        
        if not updated_user:
            return Response({
                'success': False,
                'error': 'Error al actualizar usuario'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return Response({
            'success': True,
            'data': {
                'id': updated_user.id,
                'username': updated_user.username,
                'email': updated_user.email,
                'full_name': updated_user.full_name,
                'role': updated_user.role,
                'is_active': updated_user.is_active,
            },
            'message': 'Usuario actualizado exitosamente'
        })
        
    except ValueError as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['DELETE'])
@permission_classes([IsAdminUser])
def user_delete(request, user_id):
    """
    DELETE /api/v1/users/{id}/
    Eliminar un usuario.
    """
    try:
        user_service = get_user_service()
        
        # Verificar que el usuario existe
        existing_user = user_service.get_user(user_id)
        if not existing_user:
            return Response({
                'success': False,
                'error': 'Usuario no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Eliminar usuario
        deleted = user_service.delete_user(user_id)
        
        if deleted:
            return Response({
                'success': True,
                'message': 'Usuario eliminado exitosamente'
            })
        else:
            return Response({
                'success': False,
                'error': 'Error al eliminar usuario'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def user_stats(request):
    """
    GET /api/v1/users/stats/
    Obtener estadísticas de usuarios.
    """
    try:
        user_service = get_user_service()
        
        stats = user_service.get_user_stats()
        
        return Response({
            'success': True,
            'data': {
                'total_users': stats.total_users,
                'active_users': stats.active_users,
                'inactive_users': stats.inactive_users,
                'super_admins': stats.super_admins,
                'regular_users': stats.regular_users,
            }
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
def user_login(request):
    """
    POST /api/v1/auth/login/
    Autenticar usuario y generar token.
    """
    try:
        user_service = get_user_service()
        
        # Validar datos requeridos
        if 'username' not in request.data or 'password' not in request.data:
            return Response({
                'success': False,
                'error': 'Username y password son requeridos'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Crear DTO de login
        login_request = LoginRequest(
            username=request.data['username'],
            password=request.data['password']
        )
        
        # Autenticar
        login_response = user_service.authenticate_user(login_request)
        
        if not login_response:
            return Response({
                'success': False,
                'error': 'Credenciales inválidas'
            }, status=status.HTTP_401_UNAUTHORIZED)
        
        return Response({
            'success': True,
            'data': {
                'token': login_response.token,
                'expires_in': login_response.expires_in,
                'user': {
                    'id': login_response.user.id,
                    'username': login_response.user.username,
                    'email': login_response.user.email,
                    'full_name': login_response.user.full_name,
                    'role': login_response.user.role,
                }
            },
            'message': 'Login exitoso'
        })
        
    except Exception as e:
        return Response({
            'success': False,
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@authentication_classes([JWTAuthentication])
@permission_classes([IsAuthenticated])
def user_profile(request):
    """
    GET /api/user/profile/
    Obtener perfil del usuario autenticado con JWT.
    """
    try:
        user = request.user
        
        profile_data = {
            'id': user.id,
            'username': user.username,
            'full_name': user.full_name,
            'email': user.email,
            'role': user.get_role_display(),
            'is_super_admin': user.is_super_admin(),
            'created_at': user.created_at.isoformat() if hasattr(user, 'created_at') else None,
            'assigned_solutions_count': user.assigned_solutions.filter(
                usersolutionassignment__is_active=True
            ).count() if not user.is_super_admin() else 'ALL'
        }
        
        return create_api_response(
            success=True,
            data=profile_data
        )
        
    except Exception as e:
        return handle_api_exception(e, "Error al obtener perfil del usuario")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_users_excel(request):
    """
    GET /api/v1/admin/export-users/
    Exportar lista de usuarios a archivo Excel.
    """
    try:
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        
        # Verificar que el usuario sea super admin
        if not request.user.is_super_admin():
            return create_api_response(
                success=False,
                error="No tienes permisos para exportar usuarios",
                status_code=status.HTTP_403_FORBIDDEN
            )
        
        user_service = get_user_service()
        users = user_service.get_all_users()
        
        # Crear workbook y worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Usuarios DESS"
        
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Definir headers
        headers = [
            'ID', 'Username', 'Nombre Completo', 'Email', 'Rol', 
            'Estado', 'Fecha de Creación', 'Soluciones Asignadas', 
            'Último Acceso', 'Es Super Admin'
        ]
        
        # Escribir headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos de usuarios
        for row, user in enumerate(users, 2):
            # Obtener soluciones asignadas
            if user.is_super_admin():
                solutions_count = "TODAS (Super Admin)"
            else:
                solutions_count = user.assigned_solutions.filter(
                    usersolutionassignment__is_active=True
                ).count()
            
            # Obtener último acceso si está disponible
            last_login = "N/A"
            if hasattr(user, 'last_login') and user.last_login:
                last_login = user.last_login.strftime("%Y-%m-%d %H:%M:%S")
            
            # Datos de la fila
            row_data = [
                user.id,
                user.username,
                user.full_name,
                user.email,
                user.get_role_display(),
                "Activo" if getattr(user, 'is_active', True) else "Inactivo",
                user.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(user, 'created_at') else "N/A",
                solutions_count,
                last_login,
                "Sí" if user.is_super_admin() else "No"
            ]
            
            # Escribir datos en las celdas
            for col, data in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col, value=data)
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Ajustar ancho de columnas
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Agregar información adicional
        info_row = len(users) + 3
        ws.cell(row=info_row, column=1, value=f"Reporte generado el: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=info_row + 1, column=1, value=f"Total de usuarios: {len(users)}")
        ws.cell(row=info_row + 2, column=1, value=f"Generado por: {request.user.full_name} ({request.user.username})")
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"usuarios_dess_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar workbook en response
        wb.save(response)
        
        return response
        
    except ImportError:
        return create_api_response(
            success=False,
            error="La librería openpyxl no está instalada. Ejecuta: pip install openpyxl",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        return handle_api_exception(e, "Error al exportar usuarios a Excel")
